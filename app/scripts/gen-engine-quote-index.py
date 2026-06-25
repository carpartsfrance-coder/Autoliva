#!/usr/bin/env python3
"""Génère les index JSON (occasion + reman) pour le devis instantané plaque -> moteur.

Lit les deux catalogues xlsx (hors repo, données métier) et écrit
src/data/engineQuote/{occasion,reman}.json, indexés par code moteur normalisé.
Le runtime de l'app ne lit QUE ces JSON (jamais du xlsx).

Relancer quand un catalogue change :
  OCC_XLSX=... REMAN_XLSX=... python3 scripts/gen-engine-quote-index.py
"""
import openpyxl, re, json, os
from collections import defaultdict

OCC_XLSX = os.environ.get(
    "OCC_XLSX",
    "/Users/killianbelabbes/Documents/moteur_stock_COMPLET_ovoko_AVEC-PRIX-FINAL.xlsx",
)
REMAN_XLSX = os.environ.get(
    "REMAN_XLSX",
    "/Users/killianbelabbes/Downloads/Asysum - Gamme moteurs 2026.xlsx",
)
OUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "src", "data", "engineQuote")


def norm(s):
    return re.sub(r"[^A-Z0-9]", "", str(s).upper()) if s not in (None, "") else ""


def headers(ws):
    row = next(ws.iter_rows(min_row=1, max_row=1))
    return {str(c.value).strip(): i for i, c in enumerate(row, 1) if c.value not in (None, "")}


def f_plain(x):
    try:
        return float(x)
    except (TypeError, ValueError):
        return None


def f_eur(x):  # format européen "6.370,00"
    if x is None:
        return None
    s = str(x).strip().replace(" ", "").replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


# ---------- OCCASION (stock Ovoko + prix final marge-cible) ----------
ws = openpyxl.load_workbook(OCC_XLSX, data_only=True, read_only=True)["Comparatif Ovoko"]
h = headers(ws)
c_code, c_km, c_mk, c_md = h["Code moteur"], h["Kilométrage"], h["Marque"], h["Modèle"]
c_prix = h["Prix vente final TTC (marge cible) €"]
occ_best, occ_count = {}, defaultdict(int)
for r in ws.iter_rows(min_row=2, values_only=True):
    if not r or not r[0]:
        continue
    code = norm(r[c_code - 1])
    if not code:
        continue
    prix = f_plain(r[c_prix - 1])
    if prix is None or prix <= 0:  # "à traiter (devis manuel)" -> pas d'offre occasion
        continue
    occ_count[code] += 1
    km = r[c_km - 1]
    b = occ_best.get(code)
    if b is None or prix < b["prix"]:  # on garde le moins cher pour le "à partir de"
        occ_best[code] = {
            "prix": int(round(prix)),
            "km": int(km) if isinstance(km, (int, float)) else None,
            "marque": r[c_mk - 1],
            "modele": r[c_md - 1],
        }
for code, b in occ_best.items():
    b["count"] = occ_count[code]

# ---------- RECONDITIONNÉ (Asysum, échange standard) ----------
ws2 = openpyxl.load_workbook(REMAN_XLSX, data_only=True, read_only=True)["Sheet1"]
h2 = headers(ws2)
c_desc, c_pvp, c_cons, c_disp = h2["Description"], h2["PVP"], h2["Consigne"], h2["Disponibilité"]
c_equip = h2.get("Équipement")  # col « Équipement » = pièces incluses dans le moteur fourni
pat = re.compile(r"compatible\s+(\S+)\s+\(([^)]+)\)\s*(.*)", re.I)
pat_type = re.compile(r"moteur\s+\w+\s+(neuf|recons)", re.I)
reman_best, reman_count = {}, defaultdict(int)
for r in ws2.iter_rows(min_row=2, values_only=True):
    if not r or not r[0]:
        continue
    desc = str(r[c_desc - 1] or "")
    m = pat.search(desc)
    if not m:
        continue
    code = norm(m.group(1))
    if not code:
        continue
    pvp = f_eur(r[c_pvp - 1])
    if pvp is None or pvp <= 0:
        continue
    reman_count[code] += 1
    b = reman_best.get(code)
    if b is None or pvp < b["pvp"]:  # le moins cher
        mt = pat_type.search(desc)
        reman_best[code] = {
            "pvp": int(round(pvp)),
            "consigne": int(round(f_eur(r[c_cons - 1]) or 0)),
            "dispo": str(r[c_disp - 1] or "").strip(),
            "label": m.group(3).strip(),
            "marque": m.group(2).strip(),
            "type": (mt.group(1).lower() if mt else ""),
            "equip": (str(r[c_equip - 1] or "").strip() if c_equip else ""),
        }
for code, b in reman_best.items():
    b["count"] = reman_count[code]

# ---------- ÉCRITURE ----------
os.makedirs(OUT_DIR, exist_ok=True)
with open(os.path.join(OUT_DIR, "occasion.json"), "w", encoding="utf-8") as f:
    json.dump(occ_best, f, ensure_ascii=False)
with open(os.path.join(OUT_DIR, "reman.json"), "w", encoding="utf-8") as f:
    json.dump(reman_best, f, ensure_ascii=False)

print(f"occasion: {len(occ_best)} codes | reman: {len(reman_best)} codes")
print("Contrôle sur les 10 codes de test :")
for c in ["4N14", "651901", "N62B44A", "G9U_630", "N53B30A", "N47D20C", "DRFB", "AXE", "AXD", "F1AGL411Y"]:
    nc = norm(c)
    o = occ_best.get(nc)
    rm = reman_best.get(nc)
    print(f"  {c:11} occ={'%d€' % o['prix'] if o else 'non':>7}  reman={'%d€' % rm['pvp'] if rm else 'non':>7}")
